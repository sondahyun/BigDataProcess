#!usr/bin/python3
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num = 2
F = 0
while True:
	col_midterm = sheet_ranges.cell(row = row_num, column = 3).value
	col_final = sheet_ranges.cell(row = row_num, column = 4).value
	col_homework = sheet_ranges.cell(row = row_num, column = 5).value
	col_attendance = sheet_ranges.cell(row = row_num, column = 6).value

	if col_midterm == None:
		break
	
	col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
	#col_grade = 'A'

	
	#if col_total < 40:
	#	col_grade = 'F'
	#	F += 1
	#else:
	col_grade = 'Q'

	sheet_ranges.cell(row = row_num, column = 7, value = col_total)
	sheet_ranges.cell(row = row_num, column = 8, value = col_grade)
	row_num += 1

stu_num = row_num - 2
AP = int(stu_num * 0.15)
A = int(stu_num * 0.3) - AP
BP = int(stu_num * 0.2)
B = int(stu_num * 0.4) - BP
re = stu_num - AP - A - BP - B #- F
CP = re // 2
C = re - CP

grade_num = [AP, A, BP, B, CP, C]
print(grade_num)
grade_str = ["A+", "A0", "B+", "B0", "C+", "C0"]

num = 0
n = 0
while num != stu_num:
	best = -100
	index = -1
	i = -1
	for i in range (2, row_num):
		if best < sheet_ranges.cell(row = i, column = 7).value:
			if sheet_ranges.cell(row = i, column = 8).value != 'Q':
				continue
			best = sheet_ranges.cell(row = i, column = 7).value
			best_index = i
	if grade_num[n] == 0:
		n += 1
		continue
	if sheet_ranges.cell(row = best_index, column = 8).value == 'Q':
		sheet_ranges.cell(row = best_index, column = 8, value = grade_str[n])
		grade_num[n] -= 1
		num += 1
wb.save(filename = "student.xlsx")

